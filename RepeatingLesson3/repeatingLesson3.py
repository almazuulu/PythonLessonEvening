from openpyxl import *

#Task1
def copyingFiles(filename, filename2):
    pass
    # with open(filename, 'r') as file:
    #     myText=file.read()
    # with open(filename2, 'w') as file1:
    #     file1.write(f'{myText}')

#Task3
# def findMinimumWord(filename):
#
#     with open(filename) as file:
#         text = file.read()
#
#     words = text.split()
#     minLen = len(words[0])
#     for word in words:
#         if len(word) < minLen:
#             minLen = len(word)
#
#     #print(minLen)
#     smallestWords = []
#     for word in words:
#         if len(word) == minLen:
#             smallestWords.append(word)
#
#     print(f'Самое мальнько по длине слово: {smallestWords}')

#Task6
def showingDataCoord(filename):
    pass

#Task7
"""
Примите от пользователя название города и отобразите 
данные только этого города из файла “food”
"""
def sortingByCity(filename):
    wb = load_workbook(filename, data_only=True)
    sheet = wb.worksheets[0]

    cells = sheet.rows

    cells2 = sheet.rows
    dataList = [[data.value for data in row if data.value!=None] for row in cells]

    dataList.pop(0)

    cityName = input('City you want to sort: ')
    """
    [
        [12,4,54],
        [12,4,54],
        [12,4,54],
    ]
    
    """

    sortedList = []
    for row in dataList:
        for data in row:
            if data==cityName:
                sortedList.append(row)

    for row in sortedList:
        for data in row:
            print(data, end=" ")

        print(end='\n')

def copyingData(filename):
    wb = load_workbook(filename, data_only=True)
    sheet = wb.worksheets[0]
    sheet2 = wb.create_sheet('data_copy')

    dataSet = sheet['B14:G27']

    dataList = [[data.value for data in row] for row in dataSet]
    #for row in dataSet:
        #dataList.append([data.value for data in row])
        # [for data in row:
        #     newList.append(data.value)]
        #dataList.append(newList)

        #print(end='\n')
    print(dataList)

    for data in dataList:
        sheet2.append(data)

    wb.save(filename)

def main():
    textFile = 'mytext.txt'
    textFile2 = 'mytext2.txt'
    excelFile = 'food.xlsx'

    copyingFiles(textFile, textFile2)
    #findMinimumWord(textFile)
    showingDataCoord(excelFile)
    #sortingByCity(excelFile)

    copyingData(excelFile)

if __name__ == '__main__':
    main()