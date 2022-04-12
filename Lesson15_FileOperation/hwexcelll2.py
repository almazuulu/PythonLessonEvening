from openpyxl import *
wb = load_workbook('food.xlsx')
sheetnames = wb.sheetnames
sheet2 = wb['Food_List']
cell2 = sheet2['F19':'G30']
listdata=[]
counter = 0
for row in cell2:
    for data in row:
        listdata.append(data.value)
        counter +=data.value
print(max(listdata))
print(min(listdata))
print(counter/len(listdata))
print(end="\n")
print('*'*50)
region = sheet2.iter_rows(min_row=2, min_col=2, max_row=245, max_col=2)
listdata1=[]
for row in region:
    for data in row:
        listdata1.append(data.value)
print(set(listdata1))
print('*'*50)
city = sheet2.iter_rows(min_row=2, min_col=3, max_row=245, max_col=3)
listdata2=[]
for row in city:
    for data in row:
        listdata2.append(data.value)
print(set(listdata2))
print('*'*50)
category = sheet2.iter_rows(min_row=2, min_col=4, max_row=245, max_col=4)
listdata3=[]
for row in category:
    for data in row:
        listdata3.append(data.value)
print(set(listdata3))

wb1=Workbook()
sheet5 = wb1.create_sheet('marks')
sheet7 = wb1['Sheet']
wb1.remove(sheet7)
newZaglovki =['Имя школьника', 'Математика', 'География', 'Астрономия', 'Химия','Физика']
counter = 1
for data6 in range(6):
    sheet5.cell(row=1,column=counter).value=newZaglovki[data6]
    counter += 1
wb1.save('school.xlsx')

"""
c) Примите от пользователя количество учеников, которых
необходимо добавить в данный файл, путем цикла примите
данные для этих учеников и заполните исходя из заголовка
задания b)
"""
wb3 = load_workbook('school.xlsx')
sheet9 = wb3.worksheets[0]
numberofstudent = int(input("Введите количество учеников: "))
counter = 2
student = 1
counter2 = 1
newlist=['Математика', 'География', 'Астрономия', 'Химия','Физика']
for data7 in range(numberofstudent):
    namestud = input(f"Введите имя ученика {student}:")
    sheet9.cell(row=counter, column=1).value = namestud
    sheet9.cell(row=counter, column=2).value = int(input('Математика: '))
    sheet9.cell(row=counter, column=3).value = int(input('География: '))
    sheet9.cell(row=counter, column=4).value = int(input('Астрономия: '))
    sheet9.cell(row=counter, column=5).value = int(input('Химия: '))
    sheet9.cell(row=counter, column=6).value = int(input('Физика: '))
    student += 1
    counter += 1

# for i in range(2,6):
#     grades = int(input(f"Оценки ученика по данному предмету: {sheet9[counter2].value}"))
#     sheet9.cell(row=2, column=counter2).value = namestud


wb3.save('school.xlsx')

